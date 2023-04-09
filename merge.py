import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
import numpy as np
import openpyxl

root = tk.Tk()
root.withdraw()

# 让用户选择第一个csv文件的路径
file_path1 = filedialog.askopenfilename(title="选择第一个csv文件",
                                        filetypes=[("CSV Files", "*.csv")])

# 检查用户是否选择了第一个csv文件，如果没有选择则弹出提示框，让用户选择是否重新选择文件或退出程序
while not file_path1:
    response = messagebox.askquestion("提示", "您没有选择第一个csv文件，请选择是否重新选择文件或退出程序！")
    if response == "yes":
        file_path1 = filedialog.askopenfilename(title="选择第一个csv文件",
                                                filetypes=[("CSV Files", "*.csv")])
    else:
        exit()

# 让用户选择第二个csv文件的路径
file_path2 = filedialog.askopenfilename(title="选择第二个csv文件",
                                        filetypes=[("CSV Files", "*.csv")])

# 检查用户是否选择了第二个csv文件，如果没有选择则弹出提示框，让用户选择是否重新选择文件或退出程序
while not file_path2:
    response = messagebox.askquestion("提示", "您没有选择第二个csv文件，请选择是否重新选择文件或退出程序！")
    if response == "yes":
        file_path2 = filedialog.askopenfilename(title="选择第二个csv文件",
                                                filetypes=[("CSV Files", "*.csv")])
    else:
        exit()

# 读取两个csv文件，并存储为DataFrame对象
df1 = pd.read_csv(file_path1, delimiter=';')
df2 = pd.read_csv(file_path2, delimiter=';')

# 将两个DataFrame对象根据列进行合并，并添加一列用于区分其来源
merged_df = pd.merge(df1, df2, how='outer', on=list(df1.columns), indicator=True)

# 根据合并后的“_merge”列将不同的部分标记为“left_only”或“right_only”
just_in_df1 = merged_df[merged_df['_merge'] == 'left_only'].drop(['_merge'], axis=1)
just_in_df2 = merged_df[merged_df['_merge'] == 'right_only'].drop(['_merge'], axis=1)

# 使用transpose方法将数据框翻转，让第一个文件放在左边，第二个文件放在右边
comparison_table = pd.concat([just_in_df1.set_index(list(df1.columns)).transpose(),
                              just_in_df2.set_index(list(df1.columns)).transpose()],
                             axis=1)

# 突出显示不同的部分
comparison_table = comparison_table.style.applymap(lambda x: 'background-color: #ffcccc', subset=(just_in_df1.index, just_in_df1.columns))\
                                       .applymap(lambda x: 'background-color: #ccffcc', subset=(just_in_df2.index, just_in_df2.columns))

# 输出对比结果，并进行自动换行
#print(comparison_table.render())

# 让用户选择输出Excel文件的路径和文件名
output_file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])

# 检查是否指定了输出文件路径
if not output_file_path:
    messagebox.showerror("错误", "您没有指定输出Excel文件的路径和文件名，请重新运行程序并指定输出路径！")
    exit()

# 将 DataFrame 数据输出为带有样式的 Excel 文件
with pd.ExcelWriter(output_file_path) as writer:
    comparison_table.to_excel(writer, sheet_name='Comparison Table', index=False)

# 加载 Excel 文件，并添加样式
wb = openpyxl.load_workbook(filename=output_file_path)
ws = wb['Comparison Table']

for row in range(2, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if cell.value == 'left_only':
            cell.fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor='FFFFCC')
        elif cell.value == 'right_only':
            cell.fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor='CCFFCC')

# 保存 Excel 文件
wb.save(output_file_path)
